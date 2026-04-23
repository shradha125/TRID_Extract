"""Build the TRID review matrix Excel file."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from extractions import EXTRACTIONS

records = json.load(open("/home/claude/records_clean.json"))
assert len(records) == len(EXTRACTIONS) == 119

META_COLS = ["ID", "Title", "Authors", "Year", "Venue"]
GEO_COLS = ["Country", "City", "Geography_Bucket"]
METHOD_COLS = ["Simulation_Type", "Software", "Modeling_Method", "Freight_Segment", "Data_Source"]
POLICY_COLS = ["Policy_Lever", "Key_Finding", "Stated_Limitations"]
GAP_COLS = ["ML_AI_Gap"]
ALL_COLS = META_COLS + GEO_COLS + METHOD_COLS + POLICY_COLS + GAP_COLS

GROUP_FILLS = {
    "META":   PatternFill("solid", start_color="1F4E78"),
    "GEO":    PatternFill("solid", start_color="2E75B6"),
    "METHOD": PatternFill("solid", start_color="548235"),
    "POLICY": PatternFill("solid", start_color="BF8F00"),
    "GAP":    PatternFill("solid", start_color="C65911"),
}
ZEBRA = PatternFill("solid", start_color="F2F2F2")
WHITE = PatternFill("solid", start_color="FFFFFF")

def group_of(col):
    if col in META_COLS: return "META"
    if col in GEO_COLS: return "GEO"
    if col in METHOD_COLS: return "METHOD"
    if col in POLICY_COLS: return "POLICY"
    return "GAP"

wb = Workbook()
ws = wb.active
ws.title = "Research Matrix"

thin = Side(style="thin", color="808080")
header_border = Border(top=thin, bottom=thin, left=thin, right=thin)

for ci, col in enumerate(ALL_COLS, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill = GROUP_FILLS[group_of(col)]
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = header_border

def venue_of(rec):
    s = (rec.get("serial") or "").strip()
    if s:
        return s.split("\n")[0].split("Publisher:")[0].strip()
    if rec.get("sponsor"):
        return rec["sponsor"].split("\n")[0].strip()
    return "N/A"

def first_author_et_al(authors_str):
    if not authors_str: return ""
    authors = [a.strip() for a in authors_str.split(";")]
    return authors[0] if len(authors) == 1 else f"{authors[0]} et al."

for i, (rec, ext) in enumerate(zip(records, EXTRACTIONS)):
    row = i + 2
    values = {
        "ID": i + 1, "Title": rec["title"],
        "Authors": first_author_et_al(rec["authors"]),
        "Year": rec["year"], "Venue": venue_of(rec),
        "Country": ext["Country"], "City": ext["City"],
        "Geography_Bucket": ext["Geography_Bucket"],
        "Simulation_Type": ext["Simulation_Type"],
        "Software": ext["Software"],
        "Modeling_Method": ext["Modeling_Method"],
        "Freight_Segment": ext["Freight_Segment"],
        "Data_Source": ext["Data_Source"],
        "Policy_Lever": ext["Policy_Lever"],
        "Key_Finding": ext["Key_Finding"],
        "Stated_Limitations": ext["Stated_Limitations"],
        "ML_AI_Gap": ext["ML_AI_Gap"],
    }
    zebra = ZEBRA if i % 2 == 0 else WHITE
    for ci, col in enumerate(ALL_COLS, 1):
        cell = ws.cell(row=row, column=ci, value=values[col])
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top",
                                   horizontal="center" if col in ("ID","Year") else "left")
        cell.fill = zebra
        cell.border = Border(top=Side(style="thin", color="DDDDDD"),
                             bottom=Side(style="thin", color="DDDDDD"),
                             left=Side(style="thin", color="EEEEEE"),
                             right=Side(style="thin", color="EEEEEE"))

widths = {
    "ID": 5, "Title": 45, "Authors": 22, "Year": 7, "Venue": 28,
    "Country": 15, "City": 22, "Geography_Bucket": 16,
    "Simulation_Type": 22, "Software": 18, "Modeling_Method": 38,
    "Freight_Segment": 22, "Data_Source": 28,
    "Policy_Lever": 32, "Key_Finding": 45, "Stated_Limitations": 30,
    "ML_AI_Gap": 55,
}
for ci, col in enumerate(ALL_COLS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = widths[col]

ws.row_dimensions[1].height = 32
for r in range(2, len(records) + 2):
    ws.row_dimensions[r].height = 85

ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_COLS))}{len(records)+1}"

# Summary sheet
summary = wb.create_sheet("Summary")
summary["A1"] = "TRID Urban-Freight Microsimulation Review — Summary"
summary["A1"].font = Font(bold=True, size=14, name="Arial")
summary.merge_cells("A1:D1")

from collections import Counter
geo_counts = Counter(e["Geography_Bucket"] for e in EXTRACTIONS)
country_counts = Counter(e["Country"] for e in EXTRACTIONS)

summary["A3"] = "Geography bucket"; summary["B3"] = "Count"
summary["A3"].font = Font(bold=True); summary["B3"].font = Font(bold=True)
r = 4
for k, v in sorted(geo_counts.items(), key=lambda x: -x[1]):
    summary.cell(row=r, column=1, value=k)
    summary.cell(row=r, column=2, value=v)
    r += 1
summary.cell(row=r+1, column=1, value="Total papers").font = Font(bold=True)
summary.cell(row=r+1, column=2, value=f"=SUM(B4:B{r-1})").font = Font(bold=True)

r += 3
summary.cell(row=r, column=1, value="Simulation type (normalized)").font = Font(bold=True)
summary.cell(row=r, column=2, value="Count").font = Font(bold=True)
r += 1
def norm_sim(s):
    sl = s.lower()
    if "agent" in sl or "abm" in sl or "mass-gt" in sl or "simmobility" in sl: return "Agent-based"
    if "monte carlo" in sl: return "Monte Carlo"
    if "discrete-event" in sl or "discrete event" in sl: return "Discrete-event"
    if "system dynamics" in sl: return "System dynamics"
    if "cellular automata" in sl: return "Cellular automata"
    if "microsim" in sl or "traffic micro" in sl: return "Traffic microsimulation"
    if "review" in sl: return "Review / taxonomy"
    return "Other / hybrid"

norm_sim_counts = Counter(norm_sim(e["Simulation_Type"]) for e in EXTRACTIONS)
for k, v in sorted(norm_sim_counts.items(), key=lambda x: -x[1]):
    summary.cell(row=r, column=1, value=k)
    summary.cell(row=r, column=2, value=v)
    r += 1

r += 2
summary.cell(row=r, column=1, value="Country / region (top 15)").font = Font(bold=True)
summary.cell(row=r, column=2, value="Count").font = Font(bold=True)
r += 1
for k, v in sorted(country_counts.items(), key=lambda x: -x[1])[:15]:
    summary.cell(row=r, column=1, value=k)
    summary.cell(row=r, column=2, value=v)
    r += 1

summary.column_dimensions["A"].width = 38
summary.column_dimensions["B"].width = 10

# Notes
notes = wb.create_sheet("Notes")
notes["A1"] = "Column definitions and methodology"
notes["A1"].font = Font(bold=True, size=13, name="Arial")
note_rows = [
    ("", ""),
    ("Column", "Meaning"),
    ("ID", "Sequential row number"),
    ("Title", "Paper title as extracted from TRID"),
    ("Authors", "First author + et al."),
    ("Year", "Publication year (project start year for TRID projects)"),
    ("Venue", "Journal / conference / sponsor"),
    ("Country", "Country where the study was implemented"),
    ("City", "Specific study area"),
    ("Geography_Bucket", "USA | North America (Canada) | Europe (EU) | Europe (other) | Asia | Oceania | LATAM | Africa | Middle East / Asia"),
    ("Simulation_Type", "Core simulation paradigm"),
    ("Software", "Named tool / platform, 'Not stated' if absent"),
    ("Modeling_Method", "Methodological specifics (VRP, logit, SD, optimization, etc.)"),
    ("Freight_Segment", "last-mile, long-haul, port-drayage, rural, etc."),
    ("Data_Source", "Empirical basis for the study"),
    ("Policy_Lever", "Single policy question the paper targets"),
    ("Key_Finding", "1-sentence takeaway"),
    ("Stated_Limitations", "Authors' acknowledged gaps / caveats"),
    ("ML_AI_Gap", "MODERATE-STANCE inferred ML/AI extensions: where NN, RL, GNNs, CV, or Bayesian optimization could extend the method. Inferred from stated limitations or from method characteristics (e.g., hand-tuned PSO -> Bayesian optimization surrogate; logit -> neural choice model; VRP heuristic -> learning-to-route)."),
    ("", ""),
    ("Phantom-record note", "Three records in the TRID export had no title/abstract (parser phantoms from stray '**Abstract**.' strings). Dropped. Final: 119 records from 122 entries in the two .docx files."),
    ("", ""),
    ("Reproducibility", "parse_trid.py regenerates records_clean.json; extractions.py holds research-content coding; build_excel.py assembles this workbook; extract_via_api.py is provided for rerunning extraction on new TRID exports."),
]
for ri, (a, b) in enumerate(note_rows, 2):
    notes.cell(row=ri, column=1, value=a).font = Font(bold=(ri==3), name="Arial", size=10)
    notes.cell(row=ri, column=2, value=b).font = Font(name="Arial", size=10)
    notes.cell(row=ri, column=2).alignment = Alignment(wrap_text=True, vertical="top")

notes.column_dimensions["A"].width = 22
notes.column_dimensions["B"].width = 110

wb.save("/home/claude/trid_review_matrix.xlsx")
print("Saved.")
